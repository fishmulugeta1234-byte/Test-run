import os
import re
import logging
from pathlib import Path

from docxtpl import DocxTemplate
from openpyxl import load_workbook

from telegram import Update
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    ContextTypes,
    filters,
)

# ============================================================
# CONFIG
# ============================================================

BASE_DIR = Path(__file__).resolve().parent

WORKOUT_TEMPLATE = BASE_DIR / "Simon_30Day_Workout_Blueprint_Updated.docx"
NUTRITION_TEMPLATE = BASE_DIR / "Simon_30Day_Nutrition_Blueprint-1.docx"
CALCULATOR_FILE = BASE_DIR / "Simon_Calorie_Macro_Calculator.xlsx"

OUTPUT_DIR = BASE_DIR / "outputs"
OUTPUT_DIR.mkdir(exist_ok=True)

BOT_TOKEN = os.environ.get("TELEGRAM_BOT_TOKEN")

logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO,
)

logger = logging.getLogger(__name__)


# ============================================================
# CHECK REQUIRED FILES
# ============================================================

def check_required_files():
    required = [
        WORKOUT_TEMPLATE,
        NUTRITION_TEMPLATE,
        CALCULATOR_FILE,
    ]

    missing = [
        file.name
        for file in required
        if not file.is_file()
    ]

    if missing:
        raise FileNotFoundError(
            "Missing required files:\n"
            + "\n".join(missing)
        )


# ============================================================
# HELPER FUNCTIONS
# ============================================================

def clean_filename(name):
    name = str(name).strip()

    name = re.sub(
        r'[<>:"/\\|?*]',
        "",
        name
    )

    name = re.sub(
        r"\s+",
        "_",
        name
    )

    return name or "Client"


def number(value, field):
    try:
        return float(
            str(value)
            .replace(",", "")
            .replace("kg", "")
            .replace("KG", "")
            .replace("cm", "")
            .replace("CM", "")
            .strip()
        )

    except (TypeError, ValueError):
        raise ValueError(
            f"{field} must be a number."
        )


# ============================================================
# CALORIE / MACRO CALCULATOR
# ============================================================

def calculate_macros(client):

    weight = number(
        client["weight"],
        "Weight"
    )

    height = number(
        client["height"],
        "Height"
    )

    age = number(
        client["age"],
        "Age"
    )

    gender = client["gender"].lower().strip()

    activity = (
        client["activity_level"]
        .lower()
        .strip()
    )

    goal = (
        client["goal"]
        .lower()
        .strip()
    )

    # --------------------------------------------------------
    # MIFFLIN-ST JEOR
    # --------------------------------------------------------

    if gender in (
        "male",
        "m",
        "man"
    ):

        bmr = (
            (10 * weight)
            + (6.25 * height)
            - (5 * age)
            + 5
        )

    elif gender in (
        "female",
        "f",
        "woman"
    ):

        bmr = (
            (10 * weight)
            + (6.25 * height)
            - (5 * age)
            - 161
        )

    else:

        bmr = (
            (10 * weight)
            + (6.25 * height)
            - (5 * age)
        )

    # --------------------------------------------------------
    # ACTIVITY FACTORS
    # --------------------------------------------------------

    activity_factors = {

        "sedentary": 1.20,

        "light": 1.375,

        "lightly active": 1.375,

        "moderate": 1.55,

        "moderately active": 1.55,

        "very active": 1.725,

        "extremely active": 1.90,
    }

    activity_factor = activity_factors.get(
        activity,
        1.55
    )

    # --------------------------------------------------------
    # MAINTENANCE CALORIES
    # --------------------------------------------------------

    maintenance = (
        bmr * activity_factor
    )

    # --------------------------------------------------------
    # GOAL ADJUSTMENT
    # --------------------------------------------------------

    if any(
        word in goal
        for word in (
            "fat loss",
            "weight loss",
            "lose",
            "cut"
        )
    ):

        calories = maintenance * 0.85

    elif any(
        word in goal
        for word in (
            "muscle gain",
            "weight gain",
            "gain",
            "bulk"
        )
    ):

        calories = maintenance * 1.10

    else:

        calories = maintenance

    calories = round(calories)

    # --------------------------------------------------------
    # PROTEIN
    # --------------------------------------------------------

    protein = round(
        weight * 2.0
    )

    # --------------------------------------------------------
    # FAT
    # --------------------------------------------------------

    fats = round(
        weight * 0.9
    )

    # --------------------------------------------------------
    # CARBS
    # --------------------------------------------------------

    remaining = (
        calories
        - (protein * 4)
        - (fats * 9)
    )

    carbs = round(
        max(remaining, 0) / 4
    )

    # --------------------------------------------------------
    # WATER
    # --------------------------------------------------------

    water = round(
        weight * 0.035,
        1
    )

    return {

        "calories": calories,

        "protein": protein,

        "fats": fats,

        "carbs": carbs,

        "water": water,

        "bmr": round(bmr),

        "maintenance_calories":
            round(maintenance),

        "activity_multiplier":
            activity_factor,

        "weekly_calories":
            calories * 7,
    }


# ============================================================
# CREATE EXCEL CALCULATOR COPY
# ============================================================

def create_calculator_copy(
    client,
    macros
):

    try:

        workbook = load_workbook(
            CALCULATOR_FILE
        )

        if "Bot Calculator" in workbook.sheetnames:

            sheet = workbook[
                "Bot Calculator"
            ]

        else:

            sheet = workbook.create_sheet(
                "Bot Calculator",
                0
            )

        sheet["A1"] = "Client"
        sheet["B1"] = client["name"]

        sheet["A2"] = "Weight (kg)"
        sheet["B2"] = number(
            client["weight"],
            "Weight"
        )

        sheet["A3"] = "Calories"
        sheet["B3"] = macros["calories"]

        sheet["A4"] = "Protein (g)"
        sheet["B4"] = macros["protein"]

        sheet["A5"] = "Carbs (g)"
        sheet["B5"] = macros["carbs"]

        sheet["A6"] = "Fats (g)"
        sheet["B6"] = macros["fats"]

        sheet["A7"] = "Water (L)"
        sheet["B7"] = macros["water"]

        output = (
            OUTPUT_DIR
            / f"{clean_filename(client['name'])}_Calculator.xlsx"
        )

        workbook.save(
            output
        )

        return output

    except Exception:

        logger.exception(
            "Could not create calculator copy."
        )

        return None


# ============================================================
# DOCUMENT CONTEXT
# ============================================================

def build_context(
    client,
    macros
):

    context = dict(client)

    context.update(macros)

    # Additional placeholder names
    context["client_name"] = client["name"]

    context["weight_kg"] = client["weight"]

    context["height_cm"] = client["height"]

    context["age_years"] = client["age"]

    context["calorie_target"] = (
        macros["calories"]
    )

    context["protein_target"] = (
        macros["protein"]
    )

    context["carb_target"] = (
        macros["carbs"]
    )

    context["fat_target"] = (
        macros["fats"]
    )

    context["water_target"] = (
        macros["water"]
    )

    return context


# ============================================================
# GENERATE WORKOUT DOCUMENT
# ============================================================

def generate_workout(
    client,
    macros
):

    template = DocxTemplate(
        str(WORKOUT_TEMPLATE)
    )

    template.render(
        build_context(
            client,
            macros
        )
    )

    filename = (
        f"{clean_filename(client['name'])}"
        "_Workout_Blueprint.docx"
    )

    output = (
        OUTPUT_DIR
        / filename
    )

    template.save(
        str(output)
    )

    return output


# ============================================================
# GENERATE NUTRITION DOCUMENT
# ============================================================

def generate_nutrition(
    client,
    macros
):

    template = DocxTemplate(
        str(NUTRITION_TEMPLATE)
    )

    template.render(
        build_context(
            client,
            macros
        )
    )

    filename = (
        f"{clean_filename(client['name'])}"
        "_Nutrition_Blueprint.docx"
    )

    output = (
        OUTPUT_DIR
        / filename
    )

    template.save(
        str(output)
    )

    return output


# ============================================================
# /START COMMAND
# ============================================================

async def start(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    message = (
        "SIMON ORIGIN TRANSFORMATION\n\n"

        "Please send the client assessment "
        "using this structured format:\n\n"

        "Name: Client Name\n"
        "Age: 28\n"
        "Gender: Male\n"
        "Height: 175 cm\n"
        "Weight: 82 kg\n"
        "Goal: Fat Loss\n"
        "Experience: Beginner\n"
        "Equipment: Gym\n"
        "Obstacle: Consistency\n"
        "Activity Level: Moderate\n"
        "Fitness Level: Beginner\n"
        "Restrictions: None\n"
        "Injuries: None\n"
        "Meals Per Day: 4\n"
        "Training Preference: Fat Loss\n\n"

        "Send all 15 lines together in ONE message."
    )

    await update.message.reply_text(
        message
    )


# ============================================================
# STRUCTURED ASSESSMENT PARSER
# ============================================================

def parse_structured_assessment(text):

    raw_data = {}

    # --------------------------------------------------------
    # READ EACH LINE
    # --------------------------------------------------------

    for line in text.splitlines():

        line = line.strip()

        if not line:
            continue

        if ":" not in line:
            continue

        key, value = line.split(
            ":",
            1
        )

        key = key.strip().lower()

        value = value.strip()

        raw_data[key] = value

    # --------------------------------------------------------
    # FIELD NAME ALIASES
    # --------------------------------------------------------

    aliases = {

        "name": [
            "name",
            "client name",
            "full name"
        ],

        "age": [
            "age"
        ],

        "gender": [
            "gender",
            "sex"
        ],

        "height": [
            "height",
            "height cm",
            "height(cm)"
        ],

        "weight": [
            "weight",
            "weight kg",
            "weight(kg)"
        ],

        "goal": [
            "goal",
            "main goal",
            "primary goal"
        ],

        "experience": [
            "experience",
            "training experience"
        ],

        "equipment": [
            "equipment",
            "available equipment"
        ],

        "obstacle": [
            "obstacle",
            "biggest obstacle",
            "main obstacle"
        ],

        "activity_level": [
            "activity level",
            "activity_level",
            "activity"
        ],

        "fitness_level": [
            "fitness level",
            "fitness_level"
        ],

        "restrictions": [
            "restrictions",
            "dietary restrictions",
            "diet restrictions"
        ],

        "injuries": [
            "injuries",
            "injury",
            "physical limitations"
        ],

        "meals_per_day": [
            "meals per day",
            "meals_per_day",
            "meals"
        ],

        "training_preference": [
            "training preference",
            "training_preference",
            "training preferences"
        ],
    }

    result = {}

    # --------------------------------------------------------
    # FIND VALUES
    # --------------------------------------------------------

    for field, possible_keys in aliases.items():

        found = None

        for key in possible_keys:

            if key in raw_data:

                found = raw_data[key]

                break

        result[field] = found

    # --------------------------------------------------------
    # REQUIRED FIELDS
    # --------------------------------------------------------

    required = [
        "name",
        "age",
        "gender",
        "height",
        "weight",
        "goal",
        "experience",
        "equipment",
        "obstacle",
        "activity_level",
        "fitness_level",
        "restrictions",
        "injuries",
        "meals_per_day",
        "training_preference",
    ]

    missing = [
        field
        for field in required
        if not result.get(field)
    ]

    if missing:

        return None, missing

    return result, []


# ============================================================
# HANDLE CLIENT ASSESSMENT
# ============================================================

async def handle_assessment(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):

    if not update.message:
        return

    if not update.message.text:
        return

    text = update.message.text.strip()

    try:

        # ----------------------------------------------------
        # PARSE STRUCTURED DATA
        # ----------------------------------------------------

        client, missing = parse_structured_assessment(
            text
        )

        if client is None:

            missing_text = "\n".join(
                f"- {field.replace('_', ' ').title()}"
                for field in missing
            )

            await update.message.reply_text(
                "Assessment incomplete.\n\n"
                "Missing fields:\n"
                f"{missing_text}\n\n"
                "Please send the assessment like this:\n\n"
                "Name: Client Name\n"
                "Age: 28\n"
                "Gender: Male\n"
                "Height: 175 cm\n"
                "Weight: 82 kg\n"
                "Goal: Fat Loss\n"
                "Experience: Beginner\n"
                "Equipment: Gym\n"
                "Obstacle: Consistency\n"
                "Activity Level: Moderate\n"
                "Fitness Level: Beginner\n"
                "Restrictions: None\n"
                "Injuries: None\n"
                "Meals Per Day: 4\n"
                "Training Preference: Fat Loss"
            )

            return

        # ----------------------------------------------------
        # VALIDATE NUMBERS
        # ----------------------------------------------------

        number(
            client["age"],
            "Age"
        )

        number(
            client["height"],
            "Height"
        )

        number(
            client["weight"],
            "Weight"
        )

        # ----------------------------------------------------
        # PROCESSING MESSAGE
        # ----------------------------------------------------

        processing = await update.message.reply_text(
            f"Processing {client['name']}...\n\n"
            "Calculating nutrition targets...\n"
            "Creating workout blueprint...\n"
            "Creating nutrition blueprint..."
        )

        # ----------------------------------------------------
        # CALCULATE MACROS
        # ----------------------------------------------------

        macros = calculate_macros(
            client
        )

        logger.info(
            "Macros calculated for %s: %s",
            client["name"],
            macros
        )

        # ----------------------------------------------------
        # GENERATE WORKOUT
        # ----------------------------------------------------

        workout_file = generate_workout(
            client,
            macros
        )

        # ----------------------------------------------------
        # GENERATE NUTRITION
        # ----------------------------------------------------

        nutrition_file = generate_nutrition(
            client,
            macros
        )

        # ----------------------------------------------------
        # CREATE EXCEL COPY
        # ----------------------------------------------------

        calculator_file = create_calculator_copy(
            client,
            macros
        )

        # ----------------------------------------------------
        # SEND WORKOUT
        # ----------------------------------------------------

        with open(
            workout_file,
            "rb"
        ) as file:

            await update.message.reply_document(
                document=file,
                filename=workout_file.name,
                caption=(
                    "Personalized "
                    "30-Day Workout Blueprint"
                )
            )

        # ----------------------------------------------------
        # SEND NUTRITION
        # ----------------------------------------------------

        with open(
            nutrition_file,
            "rb"
        ) as file:

            await update.message.reply_document(
                document=file,
                filename=nutrition_file.name,
                caption=(
                    "Personalized "
                    "30-Day Nutrition Blueprint"
                )
            )

        # ----------------------------------------------------
        # SEND SUMMARY
        # ----------------------------------------------------

        await update.message.reply_text(
            f"Completed for {client['name']}.\n\n"
            f"Calories: {macros['calories']} kcal/day\n"
            f"Protein: {macros['protein']} g/day\n"
            f"Carbs: {macros['carbs']} g/day\n"
            f"Fats: {macros['fats']} g/day\n"
            f"Water: {macros['water']} L/day"
        )

        # ----------------------------------------------------
        # CLEAN TEMPORARY FILES
        # ----------------------------------------------------

        for file in (
            workout_file,
            nutrition_file,
            calculator_file
        ):

            if file and file.exists():

                try:
                    file.unlink()

                except OSError:

                    logger.warning(
                        "Could not delete %s",
                        file
                    )

        # ----------------------------------------------------
        # DELETE PROCESSING MESSAGE
        # ----------------------------------------------------

        try:

            await processing.delete()

        except Exception:

            pass

    except Exception as error:

        logger.exception(
            "Assessment processing failed."
        )

        await update.message.reply_text(
            "ERROR PROCESSING ASSESSMENT\n\n"
            f"{error}\n\n"
            "Please check the assessment format "
            "and try again."
        )


# ============================================================
# TELEGRAM ERROR HANDLER
# ============================================================

async def error_handler(
    update,
    context
):

    logger.error(
        "Telegram error: %s",
        context.error,
        exc_info=context.error
    )


# ============================================================
# MAIN
# ============================================================

def main():

    # --------------------------------------------------------
    # BOT TOKEN
    # --------------------------------------------------------

    if not BOT_TOKEN:

        raise RuntimeError(
            "TELEGRAM_BOT_TOKEN is not set "
            "in Render Environment Variables."
        )

    # --------------------------------------------------------
    # CHECK FILES
    # --------------------------------------------------------

    logger.info(
        "Checking required files..."
    )

    check_required_files()

    logger.info(
        "All required files found."
    )

    # --------------------------------------------------------
    # START APPLICATION
    # --------------------------------------------------------

    logger.info(
        "Starting Simon Origin Transformation Bot..."
    )

    app = (
        Application
        .builder()
        .token(BOT_TOKEN)
        .build()
    )

    # /start
    app.add_handler(
        CommandHandler(
            "start",
            start
        )
    )

    # Structured assessment
    app.add_handler(
        MessageHandler(
            filters.TEXT & ~filters.COMMAND,
            handle_assessment
        )
    )

    # Error handler
    app.add_error_handler(
        error_handler
    )

    logger.info(
        "Bot is running."
    )

    app.run_polling()


# ============================================================
# RUN
# ============================================================

if __name__ == "__main__":
    main()
